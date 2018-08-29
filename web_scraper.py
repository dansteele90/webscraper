'''
This script should be able to load a list of websites, scrape the image and text
Transpose it into a word document and save it

Objectives:
1. Open a website
2. scrape the information from the website
3. Open a word document
4. Insert the scraped information into the Document
5. Save the Document

Alternatively:
3. Recreate the webpage locally
4. Save locally
5. Upload file to google drive

Bonus:
A. Upload the saved documents into a google drive folder
B. Open Google Drive File/Excel File, Pull out URLs from comments and create list of URLS to 
iterate over and satisfy objective 1
C. Create a database of common ingredients/recipes
D. Create a program where you can enter a list of ingredients and return recipes using those ingredients

You need to install these packages:
- Requests
- BeautifulSoup4
'''

from requests import get
from requests.exceptions import RequestException
from contextlib import closing
from bs4 import BeautifulSoup 
import html
import requests
from win32com.client import Dispatch
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
import os.path
import string
## looking at page source, we should be able to save the files as the <title> tag
## ingredient list and method contained within a json script
## need to investigate how to scrape and transform
## potential to just rip out the source and paste it into notepad as .html file 

def simple_get(url):
    """
    Attempts to get the content at `url` by making an HTTP GET request.
    If the content-type of response is some kind of HTML/XML, return the
    text content, otherwise return None.
    """
    try:
        with closing(get(url, stream=True, timeout=10)) as resp:
            if is_good_response(resp):
                return resp #.content
            else:
                return None

    except RequestException as e:
        log_error('Error during requests to {0} : {1}'.format(url, str(e)))
        return None


def is_good_response(resp):
    """
    Returns True if the response seems to be HTML, False otherwise.
    """
    content_type = resp.headers['Content-Type'].lower()
    return (resp.status_code == 200 
            and content_type is not None 
            and content_type.find('html') > -1)


def log_error(e):
    """
    It is always a good idea to log errors. 
    This function just prints them, but you can
    make it do anything.
    """
    print(e)


def scrape_contents(content):
    pass

def web_title(content):
    '''
    Find the title of the web page
    Strip out characters after | in order to save correctly
    '''
    soup = BeautifulSoup(content, 'html.parser')
    title = soup.title.string
    head, sep, tail = title.partition('|')
    return head

def col2num(col):
    num = 0
    for c in col:
        if c in string.ascii_letters:
            num = num * 26 + (ord(c.upper()) - ord('A')) + 1
    return num

def pull_comments_from_xls_file(file):

    top_cell = ''
    file_created_count = 0
    col_letter = ''
    workbook = load_workbook(file)
    sheet_names = workbook.sheetnames
    http_string = 'http'

    for i in sheet_names:
        worksheet = workbook[i]
        #for row in worksheet.iter_rows():
        for col in worksheet.iter_cols():
            for cell_chk in col:
                

                #top_cell = row.cell()
                #print(worksheet.cell(row=1, column=col).value)
                   # print(top_cell)
                if cell_chk.comment:

                    col_letters = str(col[0])
                    col_letters = col_letters[-3:]
                    print(set(col_letters))
                    print(col_letters[-3:-2])

                    top_cell = (worksheet.cell(row=1, column=col2num(col_letters[-3:-2])).value)
                    print(top_cell)
                    #print(cell.comment.text)
                    comm = str(cell_chk.comment)
                    #print(comm[9:])
                    
                    url, _ = comm[9:].split("\n")
                    #print(url)
                    #print(_[:4])
                    '''
                    Check that the either the url or _ begins with the http protocol
                    and therefore can be loaded and scraped
                    '''
                    if url[:4] == http_string: 
                        create_text_file_from_comments(url, i, top_cell)
                    elif _[:4] == http_string:
                        create_text_file_from_comments(_, i, top_cell)
                    else:
                        print('No Valid website in comment:\n {}'.format(comm))
                    file_created_count += 1

    print('Program Finished - total files created {}'.format(file_created_count))


def create_text_file(title, content):
    '''
    Open a new file, and save it with the title of the page scraped
    convert the content into utf-8, save and close
    '''
    with open(title+".html", "wb") as f:
        f.write(content.encode('utf-8'))
        f.close()
    
def duplicate_file_check(title_list, title):
    pass


def create_text_file_from_comments(url, ws, col_hdr):
    '''
    Open a new file, and save it with the title of the page scraped
    convert the content into utf-8, save and close
    '''
    
    print('Beginning file creation...')
    save_path = 'C:/Users/zf263xr/Documents/Web_Scraping/{0}/{1}'.format(ws, col_hdr)
    print(save_path)
    content = requests.get(url)
    title_list += [web_title(text_content)]
    print(title_list)
    if content == None:
        pass
    
    else:
        text_content = content.text
        encode_content = text_content.encode('utf-8')
        title = web_title(text_content)
        print('File title: {}. Writing into file now'.format(title))
        complete_name = os.path.join(save_path, title+".html")
        print(complete_name)
        with open(complete_name, "wb") as f:
            f.write(encode_content) #(content.encode('utf-8'))
            f.close()
            print("File Created: " + title)


pull_comments_from_xls_file(input("File Name: (If not in directory, please specify full file path) "))


### current run book
#from web_scraper import simple_get
#html_doc = simple_get('https://www.bonappetit.com/recipe/old-school-garlic-bread')
#html_doc = simple_get('https://www.bonappetit.com/recipe/ba-patty-melt')

#html_doc = simple_get('https://www.bonappetit.com/recipe/classic-spinach-salad')



## Get the url and scrape out the text element


# html_doc = requests.get('https://www.bonappetit.com/recipe/classic-spinach-salad')
# content_string = html_doc.text

# #from bs4 import BeautifulSoup
# # soup = BeautifulSoup(html_doc, 'html.parser')
# #print(web_title(html_doc))
# l_title = web_title(content_string)
# print(l_title)
# #Classic Garlic Bread Recipe | Bon Appetit
# #from web_scraper import create_text_file
# create_text_file(l_title, content_string)